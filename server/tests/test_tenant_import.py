"""
Phase 8 — bulk tenant import.

The thing that must never happen here is a half-finished import: some tenants
in, some silently missing, and an estate nobody can reconcile. So every row is
checked before anything is written, and arrears come in as real invoices rather
than a bare number on the tenant row.
"""

import io
import uuid
from decimal import Decimal

import pytest
from werkzeug.security import generate_password_hash

from extensions import db
from models import (
    Invoice, InvoiceLineItem, Landlord, LandlordSettings, Property, Tenant,
    Unit, User,
)
from services import tenant_import_service as importer


def _uniq():
    return uuid.uuid4().hex[:8]


@pytest.fixture()
def landlord(db_session):
    n = _uniq()
    user = User(
        email=f"imp-{n}@test.sahilpay", phone=f"2547{n[:7]}",
        password_hash=generate_password_hash("Testpass1"),
        role="landlord", is_verified=True, is_active=True,
    )
    db_session.add(user)
    db_session.flush()
    ll = Landlord(user_id=user.id, company_name=f"Import {n}",
                  abbreviated_name="IMP", currency="KES")
    db_session.add(ll)
    db_session.flush()
    db_session.add(LandlordSettings(landlord_id=ll.id))
    db_session.commit()
    return ll


def _sheet(rows: list[dict]) -> io.BytesIO:
    """Build an .xlsx in memory the way a landlord's filled-in template looks."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenants"
    headers = [h for _, h, _ in importer.COLUMNS]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key, "") for key, _, _ in importer.COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.filename = "import.xlsx"
    return buf


def _row(**overrides):
    base = {
        "property_name": "Sunrise", "unit_name": f"A{_uniq()[:4]}",
        "rent_amount": "12000", "first_name": "Amina", "last_name": "Otieno",
        "phone": f"+2547{uuid.uuid4().int % 10**8:08d}",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def test_template_has_every_column_and_a_notes_sheet(app):
    from openpyxl import load_workbook

    with app.app_context():
        wb = load_workbook(io.BytesIO(importer.build_template_workbook()))

    assert "Tenants" in wb.sheetnames
    assert "Notes" in wb.sheetnames, "the template must explain its own columns"

    headers = [str(c.value or "") for c in wb["Tenants"][1]]
    for _, name, required in importer.COLUMNS:
        assert any(name in h for h in headers), f"{name} missing from the template"
        if required:
            assert any(h == f"{name} *" for h in headers), (
                f"{name} is required but isn't starred in the header"
            )


# ---------------------------------------------------------------------------
# Validation — writes nothing
# ---------------------------------------------------------------------------

def test_validation_writes_nothing(app, landlord):
    before = Tenant.query.filter_by(landlord_id=landlord.id).count()
    rows, fatal = importer.parse_upload(_sheet([_row(), _row()]))
    assert not fatal

    result = importer.validate_rows(landlord.id, rows)
    assert result["summary"]["valid"] == 2
    assert Tenant.query.filter_by(landlord_id=landlord.id).count() == before, (
        "validation created rows — it must be a dry run"
    )


def test_missing_required_column_is_rejected_outright(app, landlord):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["property_name", "unit_name"])   # no tenant name or phone
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.filename = "bad.xlsx"

    rows, fatal = importer.parse_upload(buf)
    assert fatal, "a file missing required columns should be refused"
    assert "first_name" in fatal[0]


def test_row_level_errors_are_reported_per_row(app, landlord):
    rows, _ = importer.parse_upload(_sheet([
        _row(),                                   # fine
        _row(rent_amount="not a number"),
        _row(first_name=""),
        _row(opening_balance="-500"),
        _row(lease_start_date="31st January"),
    ]))
    result = importer.validate_rows(landlord.id, rows)

    assert result["summary"]["valid"] == 1
    assert result["summary"]["errors"] == 4
    problems = " ".join(e for r in result["rows"] for e in r["errors"])
    assert "rent_amount must be a number" in problems
    assert "first_name is required" in problems
    assert "cannot be negative" in problems
    assert "must be a date" in problems


def test_duplicate_account_number_inside_the_file_is_an_error(app, landlord):
    rows, _ = importer.parse_upload(_sheet([
        _row(account_number="DUP-1"),
        _row(account_number="DUP-1"),
    ]))
    result = importer.validate_rows(landlord.id, rows)
    problems = " ".join(e for r in result["rows"] for e in r["errors"])
    assert "appears twice in this file" in problems, (
        "two tenants sharing an account number would send their payments to "
        "the same ledger"
    )


def test_occupied_unit_is_an_error(app, landlord, db_session):
    prop = Property(landlord_id=landlord.id, name="Sunrise", number_of_units=1, city="Nairobi")
    db_session.add(prop)
    db_session.flush()
    unit = Unit(property_id=prop.id, name="A1", rent_amount=Decimal("12000"), is_occupied=True)
    db_session.add(unit)
    db_session.commit()

    rows, _ = importer.parse_upload(_sheet([_row(property_name="Sunrise", unit_name="A1")]))
    result = importer.validate_rows(landlord.id, rows)
    assert result["rows"][0]["ok"] is False
    assert "already has a tenant" in " ".join(result["rows"][0]["errors"])


def test_new_property_and_unit_are_warnings_not_errors(app, landlord):
    rows, _ = importer.parse_upload(_sheet([_row(property_name="Brand New Block")]))
    result = importer.validate_rows(landlord.id, rows)
    row = result["rows"][0]

    assert row["ok"] is True, "creating a property is normal, not an error"
    assert "create_property" in row["actions"]
    assert "create_unit" in row["actions"]


def test_row_limit_is_enforced(app, landlord):
    rows, fatal = importer.parse_upload(_sheet([_row() for _ in range(importer.MAX_ROWS + 1)]))
    assert fatal and "limit is" in fatal[0]


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

def test_commit_creates_the_estate(app, landlord):
    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="A1", account_number="SUN-A1"),
        _row(property_name="Sunrise", unit_name="A2", account_number="SUN-A2"),
        _row(property_name="Green Court", unit_name="G1", account_number="GRN-G1"),
    ]))
    result = importer.commit_rows(landlord, rows)

    assert result["created"]["tenants"] == 3
    assert result["created"]["properties"] == 2
    assert result["created"]["units"] == 3

    assert Tenant.query.filter_by(landlord_id=landlord.id).count() == 3
    # Every imported unit must be marked occupied, or it shows as vacant.
    units = Unit.query.join(Property).filter(Property.landlord_id == landlord.id).all()
    assert all(u.is_occupied for u in units)


def test_opening_balance_becomes_a_traceable_invoice(app, landlord):
    """
    Arrears must carry provenance. A bare number on the tenant row would show
    the right balance but appear nowhere on a statement and count for nothing in
    the payment score.
    """
    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="B1",
             account_number="SUN-B1", opening_balance="18000"),
    ]))
    importer.commit_rows(landlord, rows)

    tenant = Tenant.query.filter_by(landlord_id=landlord.id, account_number="SUN-B1").first()
    assert tenant is not None
    assert tenant.balance == Decimal("-18000.00"), "the tenant should owe 18,000"

    invoice = Invoice.query.filter_by(tenant_id=tenant.id).first()
    assert invoice is not None, "no invoice was raised for the opening balance"
    assert invoice.total_amount == Decimal("18000.00")

    line = InvoiceLineItem.query.filter_by(invoice_id=invoice.id).first()
    assert line.subcategory == "balance", (
        "opening arrears must be a rent ARREARS line so reports and the "
        "tenant score treat it as carried-forward rent"
    )
    assert line.category_id is not None


def test_advance_credit_is_recorded_in_the_ledger(app, landlord):
    from models import CreditLedger

    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="C1",
             account_number="SUN-C1", credit_balance="5000"),
    ]))
    importer.commit_rows(landlord, rows)

    tenant = Tenant.query.filter_by(landlord_id=landlord.id, account_number="SUN-C1").first()
    assert tenant.credit_balance == Decimal("5000")

    ledger = CreditLedger.query.filter_by(tenant_id=tenant.id).all()
    assert len(ledger) == 1, (
        "credit_balance must equal the sum of the credit ledger — a bare "
        "column value breaks that invariant"
    )
    assert ledger[0].amount == Decimal("5000")


def test_commit_skips_bad_rows_and_keeps_the_good_ones(app, landlord):
    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="D1", account_number="SUN-D1"),
        _row(rent_amount="oops"),
        _row(property_name="Sunrise", unit_name="D2", account_number="SUN-D2"),
    ]))
    result = importer.commit_rows(landlord, rows)

    assert result["created"]["tenants"] == 2
    assert len(result["skipped"]) == 1
    assert result["skipped"][0]["errors"]


def test_missing_account_numbers_are_generated_and_unique(app, landlord):
    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="E1"),
        _row(property_name="Sunrise", unit_name="E2"),
    ]))
    importer.commit_rows(landlord, rows)

    numbers = [
        t.account_number
        for t in Tenant.query.filter_by(landlord_id=landlord.id).all()
    ]
    assert all(numbers), "every tenant needs an account number to be paid to"
    assert len(set(numbers)) == len(numbers), "generated account numbers collided"


def test_phone_formats_are_accepted_as_given(app, landlord):
    """A landlord's sheet will have 07…, +254… and 254… all mixed together."""
    rows, _ = importer.parse_upload(_sheet([
        _row(property_name="Sunrise", unit_name="F1", phone="0712345678"),
        _row(property_name="Sunrise", unit_name="F2", phone="+254733111222"),
        _row(property_name="Sunrise", unit_name="F3", phone="254744555666"),
    ]))
    result = importer.validate_rows(landlord.id, rows)
    assert result["summary"]["valid"] == 3
