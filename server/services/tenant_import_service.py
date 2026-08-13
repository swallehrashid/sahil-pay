"""
services/tenant_import_service.py — bring an existing estate in from a spreadsheet.

Onboarding a real client means typing in what they already have: a book, an
Excel sheet, a photo of a ledger. Doing that by hand for 200 tenants is a day's
work per client and the main thing standing between "interested" and "using it".

The flow is deliberately two-step:

    validate()  parse the file, check every row, report problems — changing
                nothing at all
    commit()    re-validate server-side, then write

Validation runs again inside commit() rather than trusting what the browser
sent back. A half-finished import is far worse than a rejected one: an estate
with some tenants in and some missing, silently, is not something anyone can
reconcile later.

WHAT IT WRITES
    properties and units that don't exist yet, tenants, and — importantly —
    OPENING BALANCES. A tenant who owed 12,000 before the migration must still
    owe 12,000 afterwards, so arrears are written as a real invoice with a rent
    ARREARS line, not as a magic number on the tenant row. That way the balance
    has provenance and shows up correctly on statements and in the tenant score.
"""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

MAX_ROWS = 2000

# (key, header, required)
COLUMNS: list[tuple[str, str, bool]] = [
    ("property_name",     "property_name",     True),
    ("unit_name",         "unit_name",         True),
    ("rent_amount",       "rent_amount",       True),
    ("first_name",        "first_name",        True),
    ("last_name",         "last_name",         True),
    ("phone",             "phone",             True),
    ("email",             "email",             False),
    ("national_id",       "national_id",       False),
    ("account_number",    "account_number",    False),
    ("lease_start_date",  "lease_start_date",  False),
    ("lease_expiry_date", "lease_expiry_date", False),
    ("move_in_date",      "move_in_date",      False),
    ("deposit_amount",    "deposit_amount",    False),
    ("deposit_paid",      "deposit_paid",      False),
    ("opening_balance",   "opening_balance",   False),
    ("credit_balance",    "credit_balance",    False),
    ("notes",             "notes",             False),
]

COLUMN_HELP = {
    "property_name":     "Block or estate name. Created if it doesn't exist yet.",
    "unit_name":         "House/door number, e.g. A12. Created if it doesn't exist yet.",
    "rent_amount":       "Monthly rent for the unit, numbers only (e.g. 12000).",
    "first_name":        "Tenant's first name.",
    "last_name":         "Tenant's last name.",
    "phone":             "07…, 01… or +254… — any format, we tidy it up.",
    "email":             "Optional. Lets them get emailed invoices and receipts.",
    "national_id":       "Optional.",
    "account_number":    "Their M-Pesa account/reference. Left blank, we generate one.",
    "lease_start_date":  "YYYY-MM-DD. Optional.",
    "lease_expiry_date": "YYYY-MM-DD. Optional.",
    "move_in_date":      "YYYY-MM-DD. Used as the start of their payment history.",
    "deposit_amount":    "Deposit agreed. Optional.",
    "deposit_paid":      "Deposit actually held. Optional.",
    "opening_balance":   "What they OWE you today, as a positive number. Becomes "
                         "a rent-arrears invoice so the balance is traceable.",
    "credit_balance":    "Money they've paid in ADVANCE, as a positive number.",
    "notes":             "Anything else you want on their record.",
}

EXAMPLE_ROWS = [
    {
        "property_name": "Sunrise Apartments", "unit_name": "A1", "rent_amount": 12000,
        "first_name": "Amina", "last_name": "Otieno", "phone": "0712345678",
        "email": "amina@example.com", "account_number": "SUN-A1",
        "lease_start_date": "2025-01-01", "move_in_date": "2025-01-01",
        "deposit_amount": 12000, "deposit_paid": 12000,
        "opening_balance": 0, "credit_balance": 0, "notes": "",
    },
    {
        "property_name": "Sunrise Apartments", "unit_name": "A2", "rent_amount": 12000,
        "first_name": "Brian", "last_name": "Kamau", "phone": "+254722333444",
        "email": "", "account_number": "SUN-A2",
        "lease_start_date": "2024-06-01", "move_in_date": "2024-06-01",
        "deposit_amount": 12000, "deposit_paid": 6000,
        "opening_balance": 18000, "credit_balance": 0,
        "notes": "Owes two months rent",
    },
]


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def build_template_workbook() -> bytes:
    """The .xlsx a landlord fills in — headers, examples, and a Notes sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenants"

    headers = [header for _, header, _ in COLUMNS]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0F0246")
    for idx, (key, header, required) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Required columns are starred in the header so it's obvious in the
        # spreadsheet itself, not only in a document nobody opens.
        if required:
            cell.value = f"{header} *"
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 6, 16)

    for example in EXAMPLE_ROWS:
        ws.append([example.get(key, "") for key, _, _ in COLUMNS])

    notes = wb.create_sheet("Notes")
    notes.append(["Column", "Required", "What to put in it"])
    for cell in notes[1]:
        cell.font = Font(bold=True)
    for key, header, required in COLUMNS:
        notes.append([header, "Yes" if required else "Optional", COLUMN_HELP.get(key, "")])
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 12
    notes.column_dimensions["C"].width = 90
    for row in notes.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    notes.append([])
    notes.append(["The two example rows are there to show the format —",
                  "", "delete them before you upload."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _clean_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_").rstrip("*").strip("_")


def parse_upload(file_storage) -> tuple[list[dict], list[str]]:
    """
    Read an .xlsx or .csv into row dicts. Returns (rows, fatal_errors).

    Header matching is forgiving about case, spaces and the '*' the template
    puts on required columns — a landlord who retypes the header row by hand
    should not have their whole file rejected over a capital letter.
    """
    filename = (getattr(file_storage, "filename", "") or "").lower()
    raw = file_storage.read()
    if not raw:
        return [], ["The file is empty."]

    try:
        if filename.endswith(".csv"):
            import csv

            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            table = [list(r) for r in reader]
        else:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            ws = wb["Tenants"] if "Tenants" in wb.sheetnames else wb.worksheets[0]
            table = [list(r) for r in ws.iter_rows(values_only=True)]
    except Exception as exc:
        return [], [f"Could not read the file: {exc}"]

    table = [r for r in table if any(str(c or "").strip() for c in r)]
    if not table:
        return [], ["The file has no rows."]

    header = [_clean_header(c) for c in table[0]]
    known = {key for key, _, _ in COLUMNS}
    missing = [
        header_name for key, header_name, required in COLUMNS
        if required and key not in header
    ]
    if missing:
        return [], [f"These required columns are missing: {', '.join(missing)}."]

    rows = []
    for line_no, raw_row in enumerate(table[1:], start=2):
        row = {}
        for idx, key in enumerate(header):
            if key in known:
                value = raw_row[idx] if idx < len(raw_row) else None
                row[key] = "" if value is None else str(value).strip()
        row["_line"] = line_no
        rows.append(row)

    if len(rows) > MAX_ROWS:
        return [], [
            f"{len(rows)} rows — the limit is {MAX_ROWS} per file. "
            "Split it and import in batches."
        ]

    return rows, []


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _to_decimal(value, field):
    if value in (None, ""):
        return None, None
    try:
        return Decimal(str(value).replace(",", "").strip()), None
    except (InvalidOperation, ValueError):
        return None, f"{field} must be a number (got '{value}')."


def _to_date(value, field):
    if value in (None, ""):
        return None, None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19] if " " in text else text, fmt).date(), None
        except ValueError:
            continue
    return None, f"{field} must be a date like 2025-01-31 (got '{value}')."


def validate_rows(landlord_id: int, rows: list[dict]) -> dict:
    """
    Check every row without writing anything.

    ERRORS block the import (the row cannot be created correctly).
    WARNINGS do not (something will be created, or a duplicate is intentional) —
    one person legitimately renting several units is normal and supported.
    """
    from models import Property, Tenant, Unit
    from services.tenant_identity_service import normalise_phone

    existing_properties = {
        p.name.strip().lower(): p
        for p in Property.query.filter_by(landlord_id=landlord_id, is_deleted=False).all()
    }
    existing_units = {}
    for unit in (
        Unit.query.join(Property, Property.id == Unit.property_id)
        .filter(Property.landlord_id == landlord_id, Unit.is_deleted.is_(False))
        .all()
    ):
        existing_units[(unit.property_id, unit.name.strip().lower())] = unit

    existing_accounts = {
        (t.account_number or "").strip().lower()
        for t in Tenant.query.filter_by(landlord_id=landlord_id, is_deleted=False).all()
        if t.account_number
    }
    existing_phones = {
        normalise_phone(t.phone)
        for t in Tenant.query.filter_by(landlord_id=landlord_id, is_deleted=False).all()
    }

    seen_accounts: set[str] = set()
    seen_units: set[tuple[str, str]] = set()

    results = []
    new_properties: set[str] = set()
    new_units: set[tuple[str, str]] = set()

    for row in rows:
        errors, warnings, actions = [], [], []
        cleaned = {}

        for key, header, required in COLUMNS:
            value = (row.get(key) or "").strip()
            if required and not value:
                errors.append(f"{header} is required.")
            cleaned[key] = value

        property_key = cleaned["property_name"].strip().lower()
        unit_key = cleaned["unit_name"].strip().lower()

        rent, err = _to_decimal(cleaned["rent_amount"], "rent_amount")
        if err:
            errors.append(err)
        elif rent is not None and rent < 0:
            errors.append("rent_amount cannot be negative.")

        for money_field in ("deposit_amount", "deposit_paid", "opening_balance", "credit_balance"):
            amount, err = _to_decimal(cleaned[money_field], money_field)
            if err:
                errors.append(err)
            elif amount is not None and amount < 0:
                errors.append(f"{money_field} cannot be negative — enter it as a positive number.")

        for date_field in ("lease_start_date", "lease_expiry_date", "move_in_date"):
            _, err = _to_date(cleaned[date_field], date_field)
            if err:
                errors.append(err)

        phone = normalise_phone(cleaned["phone"])
        if cleaned["phone"] and not phone:
            errors.append(f"phone doesn't look like a number ('{cleaned['phone']}').")

        # Property / unit existence
        if property_key and property_key not in existing_properties:
            if property_key not in new_properties:
                new_properties.add(property_key)
            actions.append("create_property")
            warnings.append(f"Property '{cleaned['property_name']}' will be created.")

        if property_key and unit_key:
            prop = existing_properties.get(property_key)
            unit = existing_units.get((prop.id, unit_key)) if prop else None
            pair = (property_key, unit_key)

            if pair in seen_units:
                errors.append(
                    f"Unit '{cleaned['unit_name']}' appears more than once in this file."
                )
            seen_units.add(pair)

            if unit is None:
                new_units.add(pair)
                actions.append("create_unit")
                warnings.append(f"Unit '{cleaned['unit_name']}' will be created.")
            elif unit.is_occupied:
                errors.append(
                    f"Unit '{cleaned['unit_name']}' already has a tenant. "
                    "Move them out first, or correct the unit name."
                )

        # Account number
        account = cleaned["account_number"].strip().lower()
        if account:
            if account in existing_accounts:
                errors.append(f"Account number '{cleaned['account_number']}' is already in use.")
            if account in seen_accounts:
                errors.append(
                    f"Account number '{cleaned['account_number']}' appears twice in this file."
                )
            seen_accounts.add(account)
        else:
            actions.append("generate_account_number")
            warnings.append("No account number — one will be generated.")

        # Same person, another unit: supported, worth flagging.
        if phone and phone in existing_phones:
            actions.append("multi_unit")
            warnings.append(
                "This phone already belongs to a tenant here — they'll hold "
                "several units, each billed separately."
            )

        results.append({
            "line": row.get("_line"),
            "data": cleaned,
            "errors": errors,
            "warnings": warnings,
            "actions": actions,
            "ok": not errors,
        })

    return {
        "rows": results,
        "summary": {
            "total": len(results),
            "valid": sum(1 for r in results if r["ok"]),
            "errors": sum(1 for r in results if not r["ok"]),
            "new_properties": len(new_properties),
            "new_units": len(new_units),
        },
    }


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

def commit_rows(landlord, rows: list[dict], actor_user_id: int | None = None) -> dict:
    """
    Create everything for the VALID rows. Invalid rows are skipped and reported.

    Runs validation again first — the browser's copy is a convenience, not an
    authority, and a client that skipped the review step must not be able to
    write rows nobody checked.
    """
    from extensions import db
    from models import (
        CreditLedger, Invoice, InvoiceLineItem, InvoiceStatus, InvoiceType,
        Property, Tenant, TenantUnitHistory, Unit,
    )
    from services.category_service import rent_category_id, seed_default_categories
    from services.tenant_identity_service import link_tenant_to_user
    from utils import gen_reference

    validation = validate_rows(landlord.id, rows)
    valid = [r for r in validation["rows"] if r["ok"]]
    skipped = [
        {"line": r["line"], "errors": r["errors"]}
        for r in validation["rows"] if not r["ok"]
    ]

    # A landlord importing before they've ever used the app has no categories,
    # and the opening-balance invoice needs the Rent one.
    seed_default_categories(landlord.id)
    db.session.flush()
    rent_cat = rent_category_id(landlord.id)

    properties = {
        p.name.strip().lower(): p
        for p in Property.query.filter_by(landlord_id=landlord.id, is_deleted=False).all()
    }

    created = {"properties": 0, "units": 0, "tenants": 0, "opening_invoices": 0, "credits": 0}
    today = datetime.utcnow().date()
    sequence = Tenant.query.filter_by(landlord_id=landlord.id).count()

    for index, row in enumerate(valid, start=1):
        data = row["data"]
        property_key = data["property_name"].strip().lower()

        prop = properties.get(property_key)
        if prop is None:
            prop = Property(
                landlord_id=landlord.id,
                name=data["property_name"].strip(),
                number_of_units=0,
                # City is required by the model but is not worth blocking an
                # import over — the landlord fills it in afterwards.
                city="—",
            )
            db.session.add(prop)
            db.session.flush()
            properties[property_key] = prop
            created["properties"] += 1

        rent, _ = _to_decimal(data["rent_amount"], "rent_amount")
        unit = (
            Unit.query.filter_by(property_id=prop.id, is_deleted=False)
            .filter(db.func.lower(Unit.name) == data["unit_name"].strip().lower())
            .first()
        )
        if unit is None:
            unit = Unit(
                property_id=prop.id, name=data["unit_name"].strip(),
                rent_amount=rent or Decimal("0"), is_occupied=False,
            )
            db.session.add(unit)
            db.session.flush()
            prop.number_of_units = (prop.number_of_units or 0) + 1
            created["units"] += 1
        elif rent is not None:
            unit.rent_amount = rent

        sequence += 1
        account_number = data["account_number"].strip() or (
            f"{(landlord.abbreviated_name or 'ACC').upper()[:6]}-{sequence:05d}"
        )

        deposit_amount, _ = _to_decimal(data["deposit_amount"], "deposit_amount")
        deposit_paid, _ = _to_decimal(data["deposit_paid"], "deposit_paid")
        opening_balance, _ = _to_decimal(data["opening_balance"], "opening_balance")
        credit_balance, _ = _to_decimal(data["credit_balance"], "credit_balance")
        lease_start, _ = _to_date(data["lease_start_date"], "lease_start_date")
        lease_expiry, _ = _to_date(data["lease_expiry_date"], "lease_expiry_date")
        move_in, _ = _to_date(data["move_in_date"], "move_in_date")

        tenant = Tenant(
            landlord_id=landlord.id, unit_id=unit.id,
            first_name=data["first_name"].strip(), last_name=data["last_name"].strip(),
            phone=data["phone"].strip(),
            email=(data["email"].strip().lower() or None),
            national_id=data["national_id"].strip() or None,
            account_number=account_number,
            deposit_amount=deposit_amount, deposit_paid=deposit_paid,
            deposit_returned=Decimal("0.00"),
            lease_start_date=lease_start, lease_expiry_date=lease_expiry,
            move_in_date=move_in or lease_start,
            notes=data["notes"].strip() or None,
            balance=Decimal("0.00"), credit_balance=Decimal("0.00"),
        )
        db.session.add(tenant)
        db.session.flush()
        created["tenants"] += 1

        unit.is_occupied = True
        db.session.add(TenantUnitHistory(
            tenant_id=tenant.id, unit_id=unit.id,
            moved_in_at=tenant.move_in_date or today,
        ))

        link_tenant_to_user(tenant)

        # Arrears become a REAL invoice with a rent-arrears line, not a bare
        # number on the tenant row: the balance then has provenance, appears
        # correctly on statements, and is counted by the payment score.
        if opening_balance and opening_balance > 0:
            invoice = Invoice(
                invoice_number=gen_reference("INV"), landlord_id=landlord.id,
                tenant_id=tenant.id, unit_id=unit.id, property_id=prop.id,
                invoice_type=InvoiceType.custom.value, issue_date=today,
                status=InvoiceStatus.open.value, total_amount=opening_balance,
                amount_paid=Decimal("0.00"), balance=opening_balance,
                title="Opening balance brought forward",
            )
            db.session.add(invoice)
            db.session.flush()
            db.session.add(InvoiceLineItem(
                invoice_id=invoice.id, item="Balance brought forward",
                description="Arrears carried over when this estate was imported.",
                quantity=Decimal("1"), unit_price=opening_balance, amount=opening_balance,
                category_id=rent_cat, subcategory="balance",
                amount_paid=Decimal("0.00"), status="open",
            ))
            tenant.balance = Decimal(str(tenant.balance)) - opening_balance
            created["opening_invoices"] += 1

        # Advance payments: a ledger row keeps credit_balance equal to the sum
        # of the ledger, which the rest of the system relies on.
        if credit_balance and credit_balance > 0:
            db.session.add(CreditLedger(
                landlord_id=landlord.id, tenant_id=tenant.id,
                amount=credit_balance, memo="Advance held at import.",
            ))
            tenant.credit_balance = credit_balance
            created["credits"] += 1

        if index % 100 == 0:
            db.session.flush()

    db.session.commit()

    return {"created": created, "skipped": skipped, "summary": validation["summary"]}
