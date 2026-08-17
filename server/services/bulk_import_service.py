"""
services/bulk_import_service.py — bring properties, units and tenants in from a
spreadsheet, whatever shape the spreadsheet happens to be.

The existing tenant importer (services/tenant_import_service.py) takes ONE sheet
with fixed headers describing a property, a unit and a tenant on every row. That
works when someone fills in our template, and not otherwise: a manager arriving
with three years of records has a properties sheet, a units sheet and a tenants
sheet, with their own column names, and no appetite for retyping any of it.

So this module is built around three ideas the fixed-header importer cannot
express:

  SEPARATE ENTITIES   properties, then units across every property, then tenants
                      across every unit. Each import knows how to find the
                      parents the previous one created.

  COLUMN MAPPING      the file's headers are whatever they are. The caller sends
                      {field: header} and we read through it, so "House No.",
                      "unit", and "Door" all work without anyone editing a file.

  REUSABLE MAPPINGS   the same spreadsheet shape arrives every month, so a
                      mapping is worth saving by name and picking next time.

TWO-PHASE, ALWAYS
    validate()  parse, clean, check, report — writing nothing
    commit()    re-validate server-side, then write

commit() re-validates rather than trusting what the browser sends back. A
half-finished import is worse than a rejected one: an estate with some units in
and some silently missing is not something anyone reconciles later.

UNIQUENESS
    A unit's account number is what a tenant quotes when paying, so two units
    sharing one means money landing on the wrong lease. It is enforced per
    LANDLORD ACCOUNT (a property manager's whole book), which is the scope
    pay_code_service already uses — including retired aliases, since the
    resolver still honours those.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

MAX_ROWS = 5000

# --------------------------------------------------------------------------
# What each entity accepts
# --------------------------------------------------------------------------
# aliases are matched against the file's own headers to pre-fill the mapping, so
# the common spellings a Kenyan rent roll actually uses need no manual work.

Field = dict


def _f(key, label, *, required=False, kind="text", help="", aliases=()) -> Field:
    return {"key": key, "label": label, "required": required, "kind": kind,
            "help": help, "aliases": tuple(aliases)}


PROPERTY_FIELDS = [
    _f("name", "Property name", required=True,
       help="Block or estate name.",
       aliases=("property", "property_name", "block", "estate", "building", "site")),
    _f("city", "City / town", help="Used on statements and invoices.",
       aliases=("town", "location", "city")),
    _f("street_name", "Street", aliases=("street", "road", "address", "street_name")),
    _f("commission_rate", "Commission %", kind="decimal",
       help="What you charge the owner on rent collected, e.g. 7.5.",
       aliases=("commission", "commission_rate", "mgmt_fee", "management_fee")),
    _f("unit_prefix", "Account-number prefix", kind="text",
       help="Short code for this block's account numbers, e.g. KLC. "
            "Left blank, one is derived from the name.",
       aliases=("prefix", "code", "abbreviation", "abbr", "short_code")),
]

UNIT_FIELDS = [
    _f("property_name", "Property", required=True,
       help="Which block this unit is in. Must already exist, or be in the "
            "properties import you ran first.",
       aliases=("property", "block", "estate", "building", "property_name")),
    _f("name", "Unit name", required=True,
       help="House/door number, e.g. A12.",
       aliases=("unit", "unit_name", "house", "house_no", "house_number",
                "door", "door_no", "room", "unit_no", "number")),
    _f("rent_amount", "Rent", required=True, kind="decimal",
       help="Monthly rent, numbers only.",
       aliases=("rent", "rent_amount", "monthly_rent", "amount", "rent_per_month")),
    _f("account_number", "Account number", kind="text",
       help="What tenants quote when paying. Must be unique across your whole "
            "account. Left blank, one is generated.",
       aliases=("account", "account_number", "account_no", "pay_code", "paycode",
                "reference", "ref", "acc_no", "billing_ref")),
    # Deposit lives on the TENANT, not the unit — models.Unit has no deposit
    # column, because what was actually taken is a fact about a tenancy rather
    # than about the room. It is offered on the tenants import instead.
    _f("notes", "Notes", aliases=("note", "notes", "remarks", "comment")),
]

TENANT_FIELDS = [
    _f("unit_account_or_name", "Unit", required=True,
       help="The unit this tenant occupies — its account number, or "
            "'Property / Unit' (e.g. Kileleshwa Court / A1).",
       aliases=("unit", "unit_name", "house", "house_no", "door", "account",
                "account_number", "account_no", "pay_code", "reference")),
    _f("first_name", "First name", required=True,
       aliases=("first", "firstname", "first_name", "given_name")),
    _f("last_name", "Last name", required=True,
       aliases=("last", "surname", "lastname", "last_name", "family_name")),
    _f("phone", "Phone", required=True,
       help="07…, 01… or +254… — any format, we tidy it up.",
       aliases=("phone", "mobile", "tel", "telephone", "phone_number", "msisdn",
                "contact")),
    _f("email", "Email", aliases=("email", "e_mail", "email_address")),
    _f("national_id", "National ID", aliases=("id", "id_no", "national_id", "idnumber")),
    _f("lease_start_date", "Lease start", kind="date",
       aliases=("lease_start", "start_date", "lease_start_date", "from")),
    _f("lease_expiry_date", "Lease expiry", kind="date",
       aliases=("lease_end", "end_date", "expiry", "lease_expiry_date", "to")),
    _f("deposit_amount", "Deposit agreed", kind="decimal",
       aliases=("deposit", "deposit_amount")),
    _f("deposit_paid", "Deposit held", kind="decimal",
       aliases=("deposit_paid", "deposit_held")),
    _f("opening_balance", "Opening balance (owes)", kind="decimal",
       help="What they OWE you today, positive. Becomes a traceable arrears "
            "invoice rather than a magic number.",
       aliases=("balance", "arrears", "opening_balance", "owing", "outstanding")),
    _f("credit_balance", "Credit (paid ahead)", kind="decimal",
       aliases=("credit", "credit_balance", "advance", "prepaid")),
    _f("notes", "Notes", aliases=("note", "notes", "remarks")),
]

ENTITIES = {
    "properties": {"label": "Properties", "fields": PROPERTY_FIELDS,
                   "order": 1,
                   "hint": "Start here. Units and tenants attach to these."},
    "units":      {"label": "Units", "fields": UNIT_FIELDS,
                   "order": 2,
                   "hint": "Units across every property at once."},
    "tenants":    {"label": "Tenants", "fields": TENANT_FIELDS,
                   "order": 3,
                   "hint": "Tenants into units, across every property at once."},
}


def entity_spec(entity: str) -> dict:
    from utils import ApiError

    spec = ENTITIES.get(entity)
    if spec is None:
        raise ApiError(f"Unknown import type '{entity}'.", status=422,
                       errors={"entity": "unknown"})
    return spec


def catalogue() -> list[dict]:
    """The importable entities and their fields, for the wizard."""
    return [
        {
            "key": key,
            "label": spec["label"],
            "order": spec["order"],
            "hint": spec["hint"],
            "fields": [
                {k: v for k, v in field.items() if k != "aliases"}
                for field in spec["fields"]
            ],
        }
        for key, spec in sorted(ENTITIES.items(), key=lambda kv: kv[1]["order"])
    ]


# --------------------------------------------------------------------------
# Reading a file
# --------------------------------------------------------------------------

def _slug(value) -> str:
    """A header reduced to something comparable: lowercase, no punctuation."""
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def parse_file(file_storage, *, sheet: str | None = None) -> dict:
    """
    Read a .csv/.xlsx/.xls into {headers, rows, sheets}.

    Deliberately makes NO demand of the headers — which columns mean what is
    decided later, by the mapping. A file is only rejected here when it cannot
    be read at all, so a manager never has to edit their spreadsheet to get as
    far as the preview.
    """
    filename = (getattr(file_storage, "filename", "") or "").lower()
    raw = file_storage.read()
    if not raw:
        return {"headers": [], "rows": [], "sheets": [], "errors": ["The file is empty."]}

    sheets: list[str] = []
    try:
        if filename.endswith(".csv") or filename.endswith(".txt"):
            import csv

            text = raw.decode("utf-8-sig", errors="replace")
            # Sniff the delimiter: exports from Excel in some locales are
            # semicolon-separated, and silently reading those as one column is
            # the kind of failure that looks like "the importer is broken".
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            table = [list(r) for r in csv.reader(io.StringIO(text), dialect)]
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            sheets = list(workbook.sheetnames)
            worksheet = workbook[sheet] if sheet and sheet in sheets else workbook.worksheets[0]
            table = [list(r) for r in worksheet.iter_rows(values_only=True)]
    except Exception as exc:
        return {"headers": [], "rows": [], "sheets": sheets,
                "errors": [f"Could not read the file: {exc}"]}

    # Drop entirely blank lines — spreadsheets are full of them.
    table = [r for r in table if any(str(c or "").strip() for c in r)]
    if not table:
        return {"headers": [], "rows": [], "sheets": sheets,
                "errors": ["The file has no rows."]}

    headers = [str(c).strip() if c is not None else "" for c in table[0]]
    # Unnamed trailing columns are an artefact of the file, not data.
    while headers and not headers[-1]:
        headers.pop()

    rows = []
    for line_no, raw_row in enumerate(table[1:], start=2):
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else None
            row[header] = "" if value is None else str(value).strip()
        row["_line"] = line_no
        rows.append(row)

    errors = []
    if len(rows) > MAX_ROWS:
        errors.append(
            f"{len(rows)} rows — the limit is {MAX_ROWS} per file. "
            "Split it and import in batches."
        )

    return {"headers": headers, "rows": rows, "sheets": sheets, "errors": errors}


def suggest_mapping(entity: str, headers: list[str]) -> dict:
    """
    A first guess at {field_key: header}, from each field's known aliases.

    Only a starting point — the wizard shows it and the person corrects it. The
    value is that a file using ordinary words ("House No.", "Rent") needs no
    corrections at all, which is the difference between a feature people use and
    one they abandon on the mapping screen.
    """
    spec = entity_spec(entity)
    by_slug = {_slug(h): h for h in headers if h}
    mapping: dict[str, str] = {}
    taken: set[str] = set()

    for field in spec["fields"]:
        candidates = (field["key"],) + field["aliases"]
        for candidate in candidates:
            header = by_slug.get(_slug(candidate))
            if header and header not in taken:
                mapping[field["key"]] = header
                taken.add(header)
                break

    return mapping


# --------------------------------------------------------------------------
# Cleaning
# --------------------------------------------------------------------------

def _to_decimal(value, label):
    if value in (None, ""):
        return None, None
    # Strip currency symbols, thousands separators and stray spaces — a rent
    # roll says "KES 25,000.00" far more often than it says "25000".
    text = re.sub(r"[^0-9.\-]", "", str(value).replace(",", ""))
    if text in ("", "-", "."):
        return None, f"{label} must be a number (got '{value}')."
    try:
        return Decimal(text), None
    except (InvalidOperation, ValueError):
        return None, f"{label} must be a number (got '{value}')."


def _to_int(value, label):
    number, error = _to_decimal(value, label)
    if error:
        return None, error
    return (int(number) if number is not None else None), None


_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
                 "%d %b %Y", "%d %B %Y", "%b %d %Y", "%m/%d/%Y")


def _to_date(value, label):
    if value in (None, ""):
        return None, None
    text = str(value).strip()
    # openpyxl hands back real datetimes for date-formatted cells.
    if " " in text and len(text) >= 19:
        text = text[:19]
        try:
            return datetime.strptime(text, "%Y-%m-%d %H:%M:%S").date(), None
        except ValueError:
            pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"{label} must be a date like 2026-01-31 (got '{value}')."


def clean_value(field: Field, raw):
    """Coerce one cell according to its field's kind. Returns (value, error)."""
    label = field["label"]
    if field["kind"] == "decimal":
        return _to_decimal(raw, label)
    if field["kind"] == "int":
        return _to_int(raw, label)
    if field["kind"] == "date":
        return _to_date(raw, label)
    text = str(raw or "").strip()
    return (text or None), None


def apply_mapping(entity: str, rows: list[dict], mapping: dict) -> list[dict]:
    """
    Turn raw header-keyed rows into cleaned field-keyed rows.

    Each result carries its own errors so the preview can show a per-row reason
    rather than one blanket "the file is invalid".
    """
    spec = entity_spec(entity)
    out = []
    for row in rows:
        values, errors = {}, []
        for field in spec["fields"]:
            header = mapping.get(field["key"])
            raw = row.get(header) if header else None
            value, error = clean_value(field, raw)
            if error:
                errors.append(error)
            if field["required"] and value in (None, ""):
                errors.append(f"{field['label']} is required.")
            values[field["key"]] = value
        out.append({"_line": row.get("_line"), "values": values, "errors": errors,
                    "warnings": []})
    return out


# --------------------------------------------------------------------------
# Account numbers
# --------------------------------------------------------------------------

def canonical_phone(raw) -> str | None:
    """
    A Kenyan number in the one form the rest of the system stores: +2547XXXXXXXX.

    normalise_phone() from tenant_identity_service is deliberately a COMPARISON
    key — it keeps the last nine digits so "+254712345678", "0712345678" and
    "0712 345 678" collapse to one person. That is the right thing to match on
    and the wrong thing to store: writing "712345678" onto the tenant row leaves
    it unlike every other tenant in the database and unusable for SMS, which
    needs a dialable number.

    So: match on the nine-digit key, store the full international form.
    """
    from services.tenant_identity_service import normalise_phone

    key = normalise_phone(raw)
    if not key or len(key) < 9:
        return None
    return f"+254{key}"


def derive_prefix(property_name: str) -> str:
    """
    A short code from a property name:
        "Kileleshwa Court" -> "KC"   (initials)
        "Riverside"        -> "RIV"  (first three letters)

    Initials when the name has several words, otherwise the first three
    letters. Read aloud over the phone and typed on a feature phone keypad, so
    it stays short and alphanumeric. A landlord who wants something else sets
    `unit_prefix` on the property and that wins.
    """
    words = [w for w in re.split(r"[^A-Za-z0-9]+", property_name or "") if w]
    if not words:
        return "U"
    if len(words) >= 2:
        return "".join(w[0] for w in words[:3]).upper()
    return words[0][:3].upper()


def compose_account_number(prefix: str, unit_name: str, *, separator: str = "-") -> str:
    """{PREFIX}{sep}{UNIT}, cleaned to the pay-code alphabet."""
    unit_part = re.sub(r"[^A-Za-z0-9]+", "", str(unit_name or "")).upper()
    prefix_part = re.sub(r"[^A-Za-z0-9]+", "", str(prefix or "")).upper()
    if not prefix_part:
        return unit_part
    return f"{prefix_part}{separator}{unit_part}"


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------

def validate(landlord_id: int, entity: str, rows: list[dict], mapping: dict,
             options: dict | None = None) -> dict:
    """
    Check a mapped file without writing anything.

    Returns {rows, summary, headers_used}. Every row carries its own errors and
    warnings, because "row 42: account number KLC-A1 is already used by unit A1
    in Kileleshwa Court" is actionable and "the file is invalid" is not.

    ERRORS block the import for that row. WARNINGS do not — they are things a
    manager should see and may well intend, the clearest example being one
    tenant holding several units.
    """
    options = options or {}
    prepared = apply_mapping(entity, rows, mapping)

    if entity == "properties":
        _validate_properties(landlord_id, prepared)
    elif entity == "units":
        _validate_units(landlord_id, prepared, options)
    elif entity == "tenants":
        _validate_tenants(landlord_id, prepared)

    valid = [r for r in prepared if not r["errors"]]
    return {
        "rows": prepared,
        "summary": {
            "total": len(prepared),
            "valid": len(valid),
            "invalid": len(prepared) - len(valid),
            "warnings": sum(1 for r in prepared if r["warnings"]),
        },
    }


def _validate_properties(landlord_id: int, prepared: list[dict]) -> None:
    from extensions import db
    from models import Property

    existing = {
        (name or "").strip().lower()
        for (name,) in db.session.query(Property.name).filter(
            Property.landlord_id == landlord_id, Property.is_deleted.is_(False)
        )
    }
    seen: dict[str, int] = {}

    for row in prepared:
        name = (row["values"].get("name") or "").strip()
        if not name:
            continue
        key = name.lower()
        if key in existing:
            # Not an error: re-running an import, or adding units to a block
            # that is already on the system, is normal. Say what will happen.
            row["warnings"].append(
                f"'{name}' already exists — it will be reused, not duplicated.")
        if key in seen:
            row["errors"].append(
                f"'{name}' appears twice in this file (also line {seen[key]}).")
        else:
            seen[key] = row["_line"]


def _validate_units(landlord_id: int, prepared: list[dict], options: dict) -> None:
    """
    The important one. A unit's account number is what a tenant quotes when
    paying, so a duplicate means money landing on the wrong lease — checked
    against the whole account, and against the rest of this file.
    """
    from extensions import db
    from models import Property, Unit
    from services import pay_code_service

    auto_number = bool(options.get("auto_account_numbers"))
    separator = str(options.get("separator") or "-")[:1] or "-"

    properties = {
        (name or "").strip().lower(): (pid, name)
        for pid, name in db.session.query(Property.id, Property.name).filter(
            Property.landlord_id == landlord_id, Property.is_deleted.is_(False)
        )
    }
    prefixes = {key: derive_prefix(name) for key, (_pid, name) in properties.items()}

    # Existing unit names per property, so a re-run does not create A1 twice.
    existing_units = {
        (pid, (name or "").strip().lower())
        for pid, name in db.session.query(Unit.property_id, Unit.name)
        .join(Property, Property.id == Unit.property_id)
        .filter(Property.landlord_id == landlord_id, Unit.is_deleted.is_(False))
    }

    seen_accounts: dict[str, int] = {}
    seen_units: dict[tuple, int] = {}

    for row in prepared:
        values = row["values"]
        property_name = (values.get("property_name") or "").strip()
        unit_name = (values.get("name") or "").strip()
        key = property_name.lower()

        prop = properties.get(key)
        already_present = False
        if property_name and prop is None:
            row["errors"].append(
                f"No property called '{property_name}'. Import your properties "
                f"first, or correct the spelling.")
        elif prop and unit_name:
            if (prop[0], unit_name.lower()) in existing_units:
                already_present = True
                row["warnings"].append(
                    f"{property_name} already has a unit '{unit_name}' — it will "
                    f"be skipped, not duplicated.")
            pair = (prop[0], unit_name.lower())
            if pair in seen_units:
                row["errors"].append(
                    f"'{unit_name}' appears twice for {property_name} "
                    f"(also line {seen_units[pair]}).")
            else:
                seen_units[pair] = row["_line"]

        # Account number: taken from the file, or composed when the landlord
        # asked us to generate them.
        account = (values.get("account_number") or "").strip()
        if not account and auto_number and property_name and unit_name:
            account = compose_account_number(
                prefixes.get(key) or derive_prefix(property_name),
                unit_name, separator=separator)
            values["account_number"] = account

        if not account:
            continue

        # A unit that already exists is going to be SKIPPED, so its account
        # number is never written. Checking it anyway means re-running an import
        # reports "already in use" for a code the unit holds itself — which
        # reads as a serious clash and is in fact nothing at all. Re-running an
        # import is a normal thing to do.
        if already_present:
            continue

        try:
            account = pay_code_service.normalise(account)
        except Exception as exc:                      # ApiError from the service
            row["errors"].append(
                f"Account number '{values.get('account_number')}' is not usable: "
                f"{getattr(exc, 'message', exc)}")
            continue
        values["account_number"] = account

        if account in seen_accounts:
            row["errors"].append(
                f"Account number '{account}' is used twice in this file "
                f"(also line {seen_accounts[account]}). Every unit needs its own "
                f"— it is what tenants quote when they pay.")
        else:
            seen_accounts[account] = row["_line"]

        if not pay_code_service.is_available(landlord_id, account):
            row["errors"].append(
                f"Account number '{account}' is already in use on this account. "
                f"Pick another — two units sharing one means payments land on "
                f"the wrong lease.")


def _validate_tenants(landlord_id: int, prepared: list[dict]) -> None:
    from extensions import db
    from models import Property, Tenant, Unit
    from services import pay_code_service
    from services.tenant_identity_service import normalise_phone

    # Resolve a unit by account number, or by "Property / Unit".
    units_by_name = {}
    for uid, uname, pname in (
        db.session.query(Unit.id, Unit.name, Property.name)
        .join(Property, Property.id == Unit.property_id)
        .filter(Property.landlord_id == landlord_id, Unit.is_deleted.is_(False))
    ):
        units_by_name[f"{(pname or '').strip().lower()}/{(uname or '').strip().lower()}"] = uid
        units_by_name.setdefault((uname or "").strip().lower(), uid)

    occupied = {
        uid for (uid,) in db.session.query(Tenant.unit_id).filter(
            Tenant.landlord_id == landlord_id, Tenant.is_deleted.is_(False),
            Tenant.unit_id.isnot(None))
    }

    existing_phones = {
        normalise_phone(phone)
        for (phone,) in db.session.query(Tenant.phone).filter(
            Tenant.landlord_id == landlord_id, Tenant.is_deleted.is_(False))
        if phone
    }

    seen_units: dict[int, int] = {}
    # One person legitimately holds several units — that is precisely why unit
    # account numbers must be unique. It is still worth surfacing, because it is
    # also what a copy-paste error looks like.
    phone_lines: dict[str, list[int]] = {}

    for row in prepared:
        values = row["values"]
        reference = (values.get("unit_account_or_name") or "").strip()
        unit_id = None

        if reference:
            unit = pay_code_service.resolve_unit(landlord_id, reference)
            unit_id = getattr(unit, "id", None)
            if unit_id is None:
                lookup = reference.lower().replace(" / ", "/").replace(" /", "/").replace("/ ", "/")
                unit_id = units_by_name.get(lookup) or units_by_name.get(reference.lower())
            if unit_id is None:
                row["errors"].append(
                    f"No unit matches '{reference}'. Use its account number, or "
                    f"'Property / Unit'.")
            else:
                values["_unit_id"] = unit_id
                if unit_id in occupied:
                    row["errors"].append(
                        f"'{reference}' already has a tenant. Move them out "
                        f"first, or import into a different unit.")
                if unit_id in seen_units:
                    row["errors"].append(
                        f"Two tenants in this file are assigned to '{reference}' "
                        f"(also line {seen_units[unit_id]}).")
                else:
                    seen_units[unit_id] = row["_line"]

        phone = values.get("phone")
        if phone:
            key = normalise_phone(phone)
            stored = canonical_phone(phone)
            if not stored:
                row["errors"].append(f"'{phone}' doesn't look like a phone number.")
            else:
                # Store the dialable form; dedupe on the nine-digit key.
                values["phone"] = stored
                phone_lines.setdefault(key, []).append(row["_line"])
                if key in existing_phones:
                    row["warnings"].append(
                        "This phone number already belongs to a tenant on your "
                        "account — the same person renting another unit.")

    # Second pass: flag the multi-unit tenants now that every line is known.
    multi = {key: lines for key, lines in phone_lines.items() if len(lines) > 1}
    for row in prepared:
        key = normalise_phone(row["values"].get("phone"))
        if key in multi:
            others = [l for l in multi[key] if l != row["_line"]]
            row["warnings"].append(
                f"Same person as line(s) {', '.join(map(str, others))} — they "
                f"will hold several units, each billed separately.")


# --------------------------------------------------------------------------
# Commit
# --------------------------------------------------------------------------

def commit(landlord, entity: str, rows: list[dict], mapping: dict,
           options: dict | None = None, *, actor_user_id: int | None = None) -> dict:
    """
    Write the valid rows. Re-validates first — see the module docstring.

    Rows with errors are SKIPPED, not fatal: a manager with 400 units and three
    bad ones wants the 397 in, with a list of what to fix. Everything committed
    in one transaction so a failure part-way cannot leave half an estate behind.
    """
    from extensions import db

    options = options or {}
    result = validate(landlord.id, entity, rows, mapping, options)
    good = [r for r in result["rows"] if not r["errors"]]

    if entity == "properties":
        created, reused = _commit_properties(landlord, good)
    elif entity == "units":
        created, reused = _commit_units(landlord, good, options)
    elif entity == "tenants":
        created, reused = _commit_tenants(landlord, good)
    else:                                             # pragma: no cover
        created, reused = 0, 0

    db.session.commit()

    from services.audit_service import record_audit

    record_audit(
        actor_user_id=actor_user_id,
        landlord_id=landlord.id,
        action=f"bulk_import_{entity}",
        entity_type=entity,
        entity_id=None,
        description=(f"Bulk import: {created} {entity} created, {reused} skipped "
                     f"as already present, {len(result['rows']) - len(good)} rejected."),
    )
    db.session.commit()

    return {
        "created": created,
        "skipped": reused,
        "rejected": len(result["rows"]) - len(good),
        "rows": result["rows"],
        "summary": result["summary"],
    }


def _commit_properties(landlord, rows: list[dict]) -> tuple[int, int]:
    from extensions import db
    from models import Property

    existing = {
        (name or "").strip().lower(): pid
        for pid, name in db.session.query(Property.id, Property.name).filter(
            Property.landlord_id == landlord.id, Property.is_deleted.is_(False))
    }

    created = reused = 0
    for row in rows:
        values = row["values"]
        name = (values.get("name") or "").strip()
        if name.lower() in existing:
            reused += 1
            continue
        prop = Property(
            landlord_id=landlord.id,
            name=name,
            city=values.get("city") or "Nairobi",
            street_name=values.get("street_name"),
            commission_rate=values.get("commission_rate"),
        )
        db.session.add(prop)
        db.session.flush()
        existing[name.lower()] = prop.id
        created += 1
    return created, reused


def _commit_units(landlord, rows: list[dict], options: dict) -> tuple[int, int]:
    from extensions import db
    from models import Property, Unit
    from services import pay_code_service

    properties = {
        (name or "").strip().lower(): pid
        for pid, name in db.session.query(Property.id, Property.name).filter(
            Property.landlord_id == landlord.id, Property.is_deleted.is_(False))
    }
    existing = {
        (pid, (name or "").strip().lower())
        for pid, name in db.session.query(Unit.property_id, Unit.name)
        .join(Property, Property.id == Unit.property_id)
        .filter(Property.landlord_id == landlord.id, Unit.is_deleted.is_(False))
    }

    created = reused = 0
    for row in rows:
        values = row["values"]
        property_id = properties.get((values.get("property_name") or "").strip().lower())
        name = (values.get("name") or "").strip()
        if property_id is None or not name:
            continue
        if (property_id, name.lower()) in existing:
            reused += 1
            continue

        unit = Unit(
            landlord_id=landlord.id,
            property_id=property_id,
            name=name,
            rent_amount=values.get("rent_amount"),
            notes=values.get("notes"),
        )
        db.session.add(unit)
        db.session.flush()

        # Through pay_code_service so the account number lands in the same place
        # the payment resolver reads, with the same uniqueness rule and alias
        # handling as one typed in by hand.
        pay_code_service.assign(unit, values.get("account_number"),
                                landlord_id=landlord.id)
        existing.add((property_id, name.lower()))
        created += 1
    return created, reused


def _commit_tenants(landlord, rows: list[dict]) -> tuple[int, int]:
    from extensions import db
    from models import Tenant, Unit

    created = 0
    for row in rows:
        values = row["values"]
        unit_id = values.get("_unit_id")
        if not unit_id:
            continue

        tenant = Tenant(
            landlord_id=landlord.id,
            unit_id=unit_id,
            first_name=values.get("first_name"),
            last_name=values.get("last_name"),
            phone=values.get("phone"),
            email=values.get("email"),
            national_id=values.get("national_id"),
            lease_start_date=values.get("lease_start_date"),
            lease_expiry_date=values.get("lease_expiry_date"),
            deposit_amount=values.get("deposit_amount"),
            deposit_paid=values.get("deposit_paid"),
            credit_balance=values.get("credit_balance") or 0,
            balance=0,
            notes=values.get("notes"),
        )
        db.session.add(tenant)
        db.session.flush()

        # The unit is now let. Without this the occupancy figures on every
        # report would still read the whole imported estate as vacant.
        unit = db.session.get(Unit, unit_id)
        if unit is not None:
            unit.is_occupied = True

        # An opening balance becomes a real arrears INVOICE rather than a number
        # written onto the tenant row: the balance then has provenance, appears
        # correctly on a statement, and can be paid off through the normal
        # allocation path. Same reasoning as the original tenant importer.
        opening = values.get("opening_balance")
        if opening and opening > 0:
            _write_opening_balance(landlord, tenant, opening)

        created += 1
    return created, 0


def _write_opening_balance(landlord, tenant, amount) -> None:
    from datetime import date

    from extensions import db
    from models import Invoice, InvoiceLineItem, InvoiceStatus, InvoiceType, Unit
    from services.category_service import rent_category_id

    today = date.today()
    # invoices.property_id is NOT NULL — an invoice has to be attributable to a
    # block for every property-level report to add up.
    unit = db.session.get(Unit, tenant.unit_id)
    invoice = Invoice(
        landlord_id=landlord.id,
        tenant_id=tenant.id,
        unit_id=tenant.unit_id,
        property_id=unit.property_id if unit else None,
        invoice_number=f"IMP-{tenant.id}-{today:%Y%m}",
        invoice_type=InvoiceType.rent.value,
        title="Opening balance (imported)",
        issue_date=today,
        due_date=today,
        total_amount=amount,
        amount_paid=0,
        status=InvoiceStatus.open.value,
    )
    db.session.add(invoice)
    db.session.flush()

    db.session.add(InvoiceLineItem(
        invoice_id=invoice.id,
        category_id=rent_category_id(landlord.id),
        subcategory="balance",
        item="Balance brought forward at import",
        # quantity x unit_price = amount; a balance is one line of itself.
        quantity=1,
        unit_price=amount,
        amount=amount,
        amount_paid=0,
    ))
    tenant.balance = amount
    db.session.flush()
