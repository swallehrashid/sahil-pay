"""
SahilPay — services/report_builder.py
======================================
The single source of truth for every landlord report.

A report generator's only job is to build a `ReportDocument`: a title, some
`meta` (letterhead + period + filters), and one or more `Section`s (each a set
of `Column`s, a list of row dicts, and optional totals). From that one object
we can produce all three outputs the UI needs:

    document_to_json(doc, selected)   -> dict  (on-screen preview + column catalog)
    render_document(doc, "pdf",  sel) -> bytes (WeasyPrint, with letterhead+signature)
    render_document(doc, "excel", sel) -> bytes (openpyxl, one sheet, stacked sections)

`selected` is an optional per-section column allow-list (the column editor):
    {"tenants": ["house_no", "name", "amount_due"], ...}
When absent, each column's `default` flag decides visibility. This lets the
landlord add/remove/reorder columns per report without any generator changes.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from html import escape
from io import BytesIO

from utils import render_pdf

# ---------------------------------------------------------------------------
# Column kinds — drive both alignment and value formatting
# ---------------------------------------------------------------------------
TEXT = "text"
MONEY = "money"
NUMBER = "number"
DATE = "date"
PERCENT = "percent"

_RIGHT_ALIGNED = {MONEY, NUMBER, PERCENT}


@dataclass
class Column:
    key: str
    label: str
    kind: str = TEXT
    default: bool = True  # visible unless the column editor says otherwise

    @property
    def align(self) -> str:
        return "right" if self.kind in _RIGHT_ALIGNED else "left"


@dataclass
class Section:
    key: str
    title: str
    columns: list[Column]
    rows: list[dict]
    totals: dict = field(default_factory=dict)  # {col_key: raw value}
    # "table"    -> header row + data rows (the default)
    # "keyvalue" -> a two-column label/value block (used by the summary section)
    kind: str = "table"
    note: str | None = None
    # Optional charts derived from this section's rows. Each spec:
    #   {"key","title","type":"bar"|"line","x":<col_key>,"y":<col_key>}
    # The comparative reports use these so each metric can be plotted per period.
    charts: list[dict] = field(default_factory=list)


@dataclass
class ReportDocument:
    report_key: str
    title: str
    meta: dict
    sections: list[Section] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Value formatting
# ---------------------------------------------------------------------------


def _fmt_value(value, kind: str, currency: str) -> str:
    if value is None or value == "":
        return "" if kind in (TEXT, DATE) else _money(0, currency) if kind == MONEY else ""
    if kind == MONEY:
        return _money(value, currency)
    if kind == PERCENT:
        try:
            return f"{float(value):.1f}%"
        except (TypeError, ValueError):
            return str(value)
    if kind == NUMBER:
        try:
            f = float(value)
            return str(int(f)) if f == int(f) else f"{f:,.2f}"
        except (TypeError, ValueError):
            return str(value)
    if kind == DATE:
        if isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m-%d")
        return str(value)
    return str(value)


def _money(value, currency: str = "KES") -> str:
    try:
        return f"{currency} {float(value):,.2f}"
    except (TypeError, ValueError):
        return f"{currency} 0.00"


def _num(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _svg_bar_chart(title: str, labels: list[str], values: list[float], width: int = 520, height: int = 180) -> str:
    """
    A dependency-free inline-SVG bar chart. WeasyPrint renders SVG natively, so
    comparative reports carry their graphs straight into the PDF — no JS, no
    server-side image toolchain.
    """
    pad_l, pad_b, pad_t = 44, 26, 24
    plot_w = width - pad_l - 12
    plot_h = height - pad_b - pad_t
    vmax = max(values + [1.0])
    n = max(len(values), 1)
    gap = 6
    bar_w = max((plot_w - gap * (n - 1)) / n, 1)

    bars = []
    for i, (lab, val) in enumerate(zip(labels, values)):
        h = (val / vmax) * plot_h if vmax else 0
        x = pad_l + i * (bar_w + gap)
        y = pad_t + (plot_h - h)
        bars.append(
            f"<rect x='{x:.1f}' y='{y:.1f}' width='{bar_w:.1f}' height='{h:.1f}' fill='#200497' rx='2'/>"
        )
        bars.append(
            f"<text x='{x + bar_w / 2:.1f}' y='{pad_t + plot_h + 16:.1f}' font-size='8' "
            f"text-anchor='middle' fill='#6b6b80'>{escape(str(lab))}</text>"
        )
        bars.append(
            f"<text x='{x + bar_w / 2:.1f}' y='{y - 3:.1f}' font-size='8' text-anchor='middle' "
            f"fill='#4a4a60'>{val:,.0f}</text>"
        )

    axis = (
        f"<line x1='{pad_l}' y1='{pad_t}' x2='{pad_l}' y2='{pad_t + plot_h}' stroke='#c8c8d4'/>"
        f"<line x1='{pad_l}' y1='{pad_t + plot_h}' x2='{width - 12}' y2='{pad_t + plot_h}' stroke='#c8c8d4'/>"
    )
    return (
        f"<div style='margin-top:10px'><svg width='{width}' height='{height}' "
        f"xmlns='http://www.w3.org/2000/svg'>"
        f"<text x='{pad_l}' y='14' font-size='11' font-weight='600' fill='#200497'>{escape(title)}</text>"
        f"{axis}{''.join(bars)}</svg></div>"
    )


def _json_scalar(value):
    """JSON-safe primitive for the preview payload (Decimals/dates -> str/float)."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


# ---------------------------------------------------------------------------
# Column selection (the column editor)
# ---------------------------------------------------------------------------


def parse_column_selection(raw: str | None) -> dict[str, list[str]]:
    """
    Parse ?columns=... into {section_key: [col_key, ...]}.

    Accepts dotted `section.col` tokens; a bare `col` token targets the section
    named "_default" and is resolved against the report's first section at
    render time. Order is preserved so the landlord can reorder columns.

        columns=tenants.name,tenants.balance,summary.value
        columns=date,item,paid           (single-section report)
    """
    if not raw:
        return {}
    selection: dict[str, list[str]] = {}
    for token in raw.split(","):
        token = token.strip()
        if not token:
            continue
        if "." in token:
            section, col = token.split(".", 1)
        else:
            section, col = "_default", token
        selection.setdefault(section, []).append(col)
    return selection


def _visible_columns(section: Section, selection: dict[str, list[str]]) -> list[Column]:
    """Resolve which columns to show, in order, honoring the column editor."""
    by_key = {c.key: c for c in section.columns}
    chosen = selection.get(section.key) or selection.get("_default")
    if chosen:
        # keep only known keys, in the caller's order
        cols = [by_key[k] for k in chosen if k in by_key]
        if cols:
            return cols
    return [c for c in section.columns if c.default]


# ---------------------------------------------------------------------------
# JSON (preview) serialization
# ---------------------------------------------------------------------------


def document_to_json(doc: ReportDocument, selection: dict[str, list[str]] | None = None) -> dict:
    selection = selection or {}
    sections_out = []
    for sec in doc.sections:
        visible = _visible_columns(sec, selection)
        visible_keys = {c.key for c in visible}
        sections_out.append(
            {
                "key": sec.key,
                "title": sec.title,
                "kind": sec.kind,
                "note": sec.note,
                # full catalog so the UI can offer add/remove; `visible` marks current state
                "columns": [
                    {
                        "key": c.key,
                        "label": c.label,
                        "kind": c.kind,
                        "align": c.align,
                        "visible": (c.key in visible_keys) if selection else c.default,
                    }
                    for c in sec.columns
                ],
                # keyvalue rows (summary) carry label/value/display — pass them
                # through whole; table rows are projected onto their columns.
                "rows": (
                    [{k: _json_scalar(v) for k, v in row.items()} for row in sec.rows]
                    if sec.kind == "keyvalue"
                    else [{c.key: _json_scalar(row.get(c.key)) for c in sec.columns} for row in sec.rows]
                ),
                "totals": {k: _json_scalar(v) for k, v in sec.totals.items()},
                "charts": sec.charts,
            }
        )
    return {
        "report_key": doc.report_key,
        "title": doc.title,
        "meta": doc.meta,
        "sections": sections_out,
    }


# ---------------------------------------------------------------------------
# Letterhead + signature (shared by every report's PDF)
# ---------------------------------------------------------------------------


def build_meta(landlord, *, report_title: str, period: str | None = None, subject: str | None = None,
               property_name: str | None = None, extra: dict | None = None) -> dict:
    """
    Assemble the letterhead/meta block from the landlord's General settings and
    Account profile. Every report carries this so exported documents are
    official: logo + company + address at the top, signature at the bottom.
    """
    meta = {
        "report_title": report_title,
        "company_name": landlord.company_name if landlord else None,
        "abbreviated_name": getattr(landlord, "abbreviated_name", None),
        "company_address": getattr(landlord, "company_address", None),
        "logo_url": getattr(landlord, "logo_url", None),
        "signature_url": getattr(landlord, "signature_url", None),
        "currency": getattr(landlord, "currency", "KES") or "KES",
        "subject": subject,            # e.g. tenant or property the report is about
        "property_name": property_name,
        "period": period,
        "generated_at": date.today().isoformat(),
    }
    if extra:
        meta.update(extra)
    return meta


_REPORT_STYLE = """
<style>
  @page { size: A4; margin: 1.6cm 1.4cm; }
  body { font-family: -apple-system, Helvetica, Arial, sans-serif; color: #1a1a2e; font-size: 12px; }
  .letterhead { display: flex; justify-content: space-between; align-items: flex-start;
                border-bottom: 2px solid #200497; padding-bottom: 12px; margin-bottom: 6px; }
  .letterhead .brand { display: flex; gap: 12px; align-items: center; }
  .letterhead img.logo { max-height: 56px; max-width: 160px; object-fit: contain; }
  .company { font-size: 17px; font-weight: 600; }
  .muted { color: #6b6b80; }
  .doc-meta { text-align: right; font-size: 11px; color: #4a4a60; }
  .doc-title { font-size: 20px; font-weight: 300; margin: 14px 0 2px; }
  .subject { font-size: 12px; color: #4a4a60; margin-bottom: 10px; }
  h2 { font-size: 14px; font-weight: 600; margin: 22px 0 6px; color: #200497; }
  table { width: 100%; border-collapse: collapse; margin-top: 6px; }
  th, td { padding: 5px 7px; border-bottom: 1px solid #e2e2e8; font-size: 11px; }
  th { background: #f4f4f8; font-weight: 600; text-align: left; }
  td.right, th.right { text-align: right; }
  tr.total-row td { font-weight: 700; border-top: 2px solid #1a1a2e; background: #fafafe; }
  .kv td:first-child { color: #4a4a60; }
  .signature { margin-top: 40px; display: flex; justify-content: flex-end; }
  .signature .block { text-align: center; min-width: 220px; }
  .signature img { max-height: 60px; margin-bottom: 4px; }
  .signature .line { border-top: 1px solid #1a1a2e; padding-top: 4px; font-size: 11px; }
</style>
"""


def _letterhead_html(meta: dict) -> str:
    logo = (
        f"<img class='logo' src='{escape(meta['logo_url'])}' alt='logo'/>"
        if meta.get("logo_url")
        else ""
    )
    company = escape(meta.get("company_name") or "")
    address = escape(meta.get("company_address") or "")
    parts = [f"<div class='company'>{company}</div>"]
    if address:
        parts.append(f"<div class='muted'>{address}</div>")
    left = f"<div class='brand'>{logo}<div>{''.join(parts)}</div></div>"

    meta_lines = [f"<div><strong>{escape(meta.get('report_title') or 'Report')}</strong></div>"]
    if meta.get("property_name"):
        meta_lines.append(f"<div>Property: {escape(str(meta['property_name']))}</div>")
    if meta.get("period"):
        meta_lines.append(f"<div>Period: {escape(str(meta['period']))}</div>")
    meta_lines.append(f"<div>Generated: {escape(meta.get('generated_at') or '')}</div>")
    meta_lines.append(f"<div>Currency: {escape(meta.get('currency') or 'KES')}</div>")
    right = f"<div class='doc-meta'>{''.join(meta_lines)}</div>"

    subject = (
        f"<div class='subject'>{escape(str(meta['subject']))}</div>" if meta.get("subject") else ""
    )
    return (
        f"<div class='letterhead'>{left}{right}</div>"
        f"<div class='doc-title'>{escape(meta.get('report_title') or 'Report')}</div>{subject}"
    )


def _signature_html(meta: dict) -> str:
    img = (
        f"<img src='{escape(meta['signature_url'])}' alt='signature'/>"
        if meta.get("signature_url")
        else ""
    )
    name = escape(meta.get("company_name") or "")
    return (
        "<div class='signature'><div class='block'>"
        f"{img}"
        f"<div class='line'>Authorised signature — {name}</div>"
        "</div></div>"
    )


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------


def render_document(doc: ReportDocument, fmt: str, selection: dict[str, list[str]] | None = None,
                    chart_keys: list[str] | None = None) -> bytes:
    if fmt == "excel":
        return _render_excel(doc, selection or {})
    return _render_pdf(doc, selection or {}, chart_keys or [])


def _render_pdf(doc: ReportDocument, selection: dict[str, list[str]], chart_keys: list[str]) -> bytes:
    currency = doc.meta.get("currency", "KES")
    body = [_letterhead_html(doc.meta)]

    for sec in doc.sections:
        body.append(f"<h2>{escape(sec.title)}</h2>")
        if sec.note:
            body.append(f"<div class='muted' style='font-size:11px'>{escape(sec.note)}</div>")

        if sec.kind == "keyvalue":
            rows_html = "".join(
                f"<tr><td>{escape(str(r.get('label', '')))}</td>"
                f"<td class='right'>{escape(str(r.get('display', r.get('value', ''))))}</td></tr>"
                for r in sec.rows
            )
            body.append(f"<table class='kv'><tbody>{rows_html}</tbody></table>")
            continue

        visible = _visible_columns(sec, selection)
        if not visible:
            continue
        head = "".join(
            f"<th class='{c.align}'>{escape(c.label)}</th>" for c in visible
        )
        rows_html = ""
        for row in sec.rows:
            cells = "".join(
                f"<td class='{c.align}'>{escape(_fmt_value(row.get(c.key), c.kind, currency))}</td>"
                for c in visible
            )
            rows_html += f"<tr>{cells}</tr>"
        if sec.totals:
            total_cells = ""
            for i, c in enumerate(visible):
                if c.key in sec.totals:
                    total_cells += f"<td class='{c.align}'>{escape(_fmt_value(sec.totals[c.key], c.kind, currency))}</td>"
                elif i == 0:
                    total_cells += "<td>Total</td>"
                else:
                    total_cells += "<td></td>"
            rows_html += f"<tr class='total-row'>{total_cells}</tr>"
        body.append(f"<table><thead><tr>{head}</tr></thead><tbody>{rows_html}</tbody></table>")

        # Charts selected for download, drawn from this section's rows.
        for chart in sec.charts:
            if chart_keys and chart["key"] not in chart_keys:
                continue
            labels = [str(r.get(chart["x"], "")) for r in sec.rows]
            values = [_num(r.get(chart["y"])) for r in sec.rows]
            if values:
                body.append(_svg_bar_chart(chart.get("title", chart["key"]), labels, values))

    body.append(_signature_html(doc.meta))
    html = (
        f"<!doctype html><html><head><meta charset='utf-8'>{_REPORT_STYLE}</head>"
        f"<body>{''.join(body)}</body></html>"
    )
    return render_pdf(html)


def _render_excel(doc: ReportDocument, selection: dict[str, list[str]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    currency = doc.meta.get("currency", "KES")
    wb = Workbook()
    ws = wb.active
    ws.title = (doc.meta.get("report_title") or "Report")[:31]

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="200497")
    header_fill = PatternFill("solid", fgColor="F4F4F8")

    r = 1
    # letterhead
    ws.cell(row=r, column=1, value=doc.meta.get("company_name") or "").font = title_font
    r += 1
    if doc.meta.get("company_address"):
        ws.cell(row=r, column=1, value=doc.meta["company_address"])
        r += 1
    ws.cell(row=r, column=1, value=doc.meta.get("report_title") or "Report").font = bold
    r += 1
    for label in ("subject", "property_name", "period"):
        if doc.meta.get(label):
            ws.cell(row=r, column=1, value=f"{label.replace('_', ' ').title()}: {doc.meta[label]}")
            r += 1
    ws.cell(row=r, column=1, value=f"Generated: {doc.meta.get('generated_at', '')}  |  Currency: {currency}")
    r += 2

    for sec in doc.sections:
        ws.cell(row=r, column=1, value=sec.title).font = Font(bold=True, size=12, color="200497")
        r += 1

        if sec.kind == "keyvalue":
            for row in sec.rows:
                ws.cell(row=r, column=1, value=str(row.get("label", "")))
                ws.cell(row=r, column=2, value=row.get("display", row.get("value", "")))
                r += 1
            r += 1
            continue

        visible = _visible_columns(sec, selection)
        if not visible:
            continue
        for ci, c in enumerate(visible, start=1):
            cell = ws.cell(row=r, column=ci, value=c.label)
            cell.font = bold
            cell.fill = header_fill
        r += 1
        for row in sec.rows:
            for ci, c in enumerate(visible, start=1):
                ws.cell(row=r, column=ci, value=_excel_value(row.get(c.key), c.kind))
            r += 1
        if sec.totals:
            ws.cell(row=r, column=1, value="Total").font = bold
            for ci, c in enumerate(visible, start=1):
                if c.key in sec.totals:
                    cell = ws.cell(row=r, column=ci, value=_excel_value(sec.totals[c.key], c.kind))
                    cell.font = bold
            r += 1
        r += 1

    # auto-width
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 12), 42)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_value(value, kind: str):
    """Numbers stay numeric in Excel; dates become ISO strings; text passes through."""
    if value is None or value == "":
        return 0 if kind in (MONEY, NUMBER, PERCENT) else ""
    if kind in (MONEY, NUMBER, PERCENT):
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    if kind == DATE and isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)
